from openpyxl import load_workbook
from mfiles_cli.xlsx import write_rent_roll, validate_xlsx


def test_xlsx_two_sheets(tmp_path):
    out = tmp_path / "zb15.xlsx"
    write_rent_roll(out, [{"Objekt": "ZB15", "Einheit": "EG", "Stimmt M-Files mit Vertrag/Beleg?": "OFFEN"}])
    validate_xlsx(out)
    wb = load_workbook(out)
    assert wb.sheetnames == ["Uebersicht", "Legende"]
    assert wb["Uebersicht"]["A1"].value == "Objekt"
