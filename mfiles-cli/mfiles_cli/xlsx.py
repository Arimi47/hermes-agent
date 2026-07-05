from pathlib import Path
import zipfile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUMNS = [
    "Objekt",
    "Einheit",
    "Mieter",
    "Unit ID",
    "Kategorie",
    "M-Files NKM aktuell",
    "Miete laut Vertrag/letztem Beleg",
    "Delta",
    "Stimmt M-Files mit Vertrag/Beleg?",
    "Letzte Erhoehung / Basisdatum",
    "Was ist das Datum?",
    "Quelle",
    "Klartext fuer Dritte",
]


def write_rent_roll(path: str | Path, rows: list[dict]) -> Path:
    out = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Uebersicht"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for idx, col in enumerate(COLUMNS, 1):
        width = min(max(len(col), 12), 42)
        ws.column_dimensions[get_column_letter(idx)].width = width
    legend = wb.create_sheet("Legende")
    legend.append(["Feld", "Bedeutung"])
    legend.append(["JA", "M-Files entspricht Vertrag/letztem Beleg."])
    legend.append(["NEIN", "M-Files weicht von Vertrag/letztem Beleg ab."])
    legend.append(["OFFEN", "Kein belastbarer Beleg gefunden oder Index/VPI muss manuell gerechnet werden."])
    legend.append(["Kategorie", "BGB, Index, Gewerbe oder Sonderfall."])
    for cell in legend[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    legend.column_dimensions["A"].width = 28
    legend.column_dimensions["B"].width = 100
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    validate_xlsx(out)
    return out


def validate_xlsx(path: str | Path) -> None:
    p = Path(path)
    assert zipfile.ZipFile(p).testzip() is None
    wb = load_workbook(p, data_only=True)
    assert wb.sheetnames == ["Uebersicht", "Legende"]
